#!/usr/bin/env python3
"""Convierte results.csv del google-maps-scraper a un .xlsx listo para Google Sheets.

- FILTRA: se queda solo con peluquerías caninas reales; descarta tiendas, bazares,
  peluquerías/barberías humanas, supermercados, etc.
- Deduplica negocios (por cid / teléfono / nombre+dirección)
- Saca la CIUDAD de la dirección real (no del UUID de la búsqueda)
- Ordena: primero los SIN web (leads calientes), luego por nº de reseñas (más activos)
- Añade columnas de tracking: Estado / Fecha contacto / Fecha demo / Notas
- Hoja 2: Veterinarios (lista plan B / upsell)
Uso: python3 convert_to_xlsx.py [results.csv] [salida.xlsx]
"""
import csv
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SRC = Path(sys.argv[1] if len(sys.argv) > 1 else Path(__file__).parent / "results.csv")
DST = Path(sys.argv[2] if len(sys.argv) > 2 else Path(__file__).parent / "leads-peluquerias-caninas.xlsx")

# Categorías que SÍ son peluquería canina / servicios caninos con citas recurrentes
GROOMER_CATS = {
    "peluquero de mascotas",
    "servicio de aseo de mascotas",
    "guardería canina",
    "cuidador de mascotas",
    "adiestrador canino",
}
# Categorías veterinarias -> hoja plan B
VET_CATS = {"veterinario", "hospital veterinario", "clínica veterinaria"}
# Categorías ambiguas: se aceptan SOLO si el nombre delata que es canina
AMBIGUOUS_CATS = {"peluquería", "estética", "spa", ""}
# Palabras en el NOMBRE que confirman que es canina
CANINE_HINTS = ("canin", "mascota", "perro", "perru", "can ", "pet", "dog", "grooming",
                "peludo", "guau", "woof", "animal", "pelu can")

COLS = [
    ("title", "Negocio"),
    ("phone", "Teléfono"),
    ("website", "Web"),
    ("_city", "Ciudad"),
    ("address", "Dirección"),
    ("category", "Categoría"),
    ("review_rating", "Rating"),
    ("review_count", "Nº reseñas"),
    ("link", "Ficha Google Maps"),
]
EXTRA = ["Sin web", "Estado llamada", "Fecha último contacto", "Fecha demo", "Notas"]


def city_of(r):
    """Ciudad desde complete_address (JSON) con fallback a un trozo de address."""
    raw = r.get("complete_address", "")
    try:
        d = json.loads(raw) if raw and raw not in ("{}", "null") else {}
        if d.get("city"):
            return d["city"]
    except (ValueError, TypeError):
        pass
    # fallback: penúltimo trozo de la dirección "..., 28036 Madrid"
    parts = [p.strip() for p in r.get("address", "").split(",") if p.strip()]
    return parts[-1] if parts else ""


def classify(r):
    cat = (r.get("category") or "").strip().lower()
    name = (r.get("title") or "").lower()
    if cat in GROOMER_CATS:
        return "groomer"
    if cat in VET_CATS:
        return "vet"
    if cat in AMBIGUOUS_CATS and any(h in name for h in CANINE_HINTS):
        return "groomer"
    return "skip"


groomers, vets, seen = [], [], set()
with open(SRC, newline="", encoding="utf-8") as f:
    for r in csv.DictReader(f):
        key = (r.get("cid") or r.get("phone") or (r.get("title", "") + r.get("address", ""))).strip().lower()
        if not key or key in seen:
            continue
        seen.add(key)
        r["_city"] = city_of(r)
        kind = classify(r)
        if kind == "groomer":
            groomers.append(r)
        elif kind == "vet":
            vets.append(r)


def sort_key(r):
    has_web = bool(r.get("website", "").strip())
    try:
        reviews = int(float(r.get("review_count") or 0))
    except ValueError:
        reviews = 0
    return (has_web, -reviews)


groomers.sort(key=sort_key)
vets.sort(key=sort_key)

header_fill = PatternFill("solid", fgColor="1F2937")
header_font = Font(color="FFFFFF", bold=True)
hot_fill = PatternFill("solid", fgColor="FEF3C7")  # amarillo: sin web = leads calientes
widths = [32, 16, 28, 16, 42, 22, 7, 10, 45, 8, 16, 18, 14, 40]


def build_sheet(ws, data):
    headers = [h for _, h in COLS] + EXTRA
    ws.append(headers)
    for c in ws[1]:
        c.fill, c.font = header_fill, header_font
        c.alignment = Alignment(vertical="center")
    for r in data:
        sin_web = "SÍ" if not r.get("website", "").strip() else ""
        row = [r.get(k, "") for k, _ in COLS] + [sin_web, "", "", "", ""]
        ws.append(row)
        if sin_web:
            for cell in ws[ws.max_row][: len(headers)]:
                cell.fill = hot_fill
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


wb = Workbook()
ws1 = wb.active
ws1.title = "Peluquerías caninas"
build_sheet(ws1, groomers)
ws2 = wb.create_sheet("Veterinarios (plan B)")
build_sheet(ws2, vets)
wb.save(DST)

g_noweb = sum(1 for r in groomers if not r.get("website", "").strip())
g_phone = sum(1 for r in groomers if r.get("phone", "").strip())
print(f"OK → {DST.name}")
print(f"  🐕 HOJA 1 · Peluquerías caninas: {len(groomers)} leads")
print(f"       🔥 {g_noweb} SIN web (amarillo, llamar primero) · 📞 {g_phone} con teléfono")
print(f"  🩺 HOJA 2 · Veterinarios (plan B): {len(vets)} leads")
